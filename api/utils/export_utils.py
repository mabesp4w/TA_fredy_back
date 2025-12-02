# utils/export_utils.py
import os
import io
import logging
from datetime import datetime
from django.http import HttpResponse
from django.conf import settings
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, Image as RLImage, KeepTogether
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from PIL import Image as PILImage
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)


class PDFExporter:
    """
    Utility class untuk export data ke PDF menggunakan ReportLab
    """
    
    def __init__(self):
        self.styles = getSampleStyleSheet()
        self._setup_custom_styles()
    
    def _setup_custom_styles(self):
        """Setup custom styles untuk PDF"""
        # Title style
        self.title_style = ParagraphStyle(
            'CustomTitle',
            parent=self.styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30,
            alignment=1  # Center alignment
        )
        
        # Heading style
        self.heading_style = ParagraphStyle(
            'CustomHeading',
            parent=self.styles['Heading2'],
            fontSize=14,
            textColor=colors.HexColor('#2c3e50'),
            spaceAfter=12
        )
        
        # Normal style
        self.normal_style = self.styles['Normal']
    
    def export_to_pdf(self, data, filename=None, title="Report", headers=None):
        """
        Export data ke PDF
        
        Args:
            data: List of dictionaries atau list of lists
            filename: Nama file (optional)
            title: Judul report
            headers: List header untuk tabel (optional)
        
        Returns:
            HttpResponse dengan PDF content
        """
        try:
            # Create response
            response = HttpResponse(content_type='application/pdf')
            
            if filename:
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
            else:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                response['Content-Disposition'] = f'attachment; filename="report_{timestamp}.pdf"'
            
            # Create PDF document
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(buffer, pagesize=A4)
            story = []
            
            # Add title
            story.append(Paragraph(title, self.title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Add date
            date_str = datetime.now().strftime('%d %B %Y, %H:%M:%S')
            story.append(Paragraph(f"Generated on: {date_str}", self.normal_style))
            story.append(Spacer(1, 0.3*inch))
            
            # Prepare table data
            table_data = []
            
            # Add headers if provided
            if headers:
                table_data.append(headers)
            
            # Add data rows
            if data:
                if isinstance(data[0], dict):
                    # Data is list of dictionaries
                    if not headers:
                        # Auto-generate headers from first dict keys
                        headers = list(data[0].keys())
                        table_data.append(headers)
                    
                    for row in data:
                        table_data.append([str(row.get(h, '')) for h in headers])
                else:
                    # Data is list of lists
                    table_data.extend(data)
            
            # Create table
            if table_data:
                table = Table(table_data)
                
                # Style the table
                table.setStyle(TableStyle([
                    # Header row
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#34495e')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 12),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    
                    # Data rows
                    ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
                    ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 10),
                    ('GRID', (0, 0), (-1, -1), 1, colors.grey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    
                    # Alternating row colors
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
                ]))
                
                story.append(table)
            else:
                story.append(Paragraph("No data available", self.normal_style))
            
            # Build PDF
            doc.build(story)
            
            # Get PDF content
            pdf = buffer.getvalue()
            buffer.close()
            
            response.write(pdf)
            return response
            
        except Exception as e:
            logger.error(f"Error generating PDF: {e}")
            raise
    
    def export_to_pdf_cards(self, data, filename=None, title="Report", image_field='image_path'):
        """
        Export data ke PDF dalam format card dengan gambar
        
        Args:
            data: List of dictionaries dengan field data dan image_path
            filename: Nama file (optional)
            title: Judul report
            image_field: Nama field yang berisi path gambar
        
        Returns:
            HttpResponse dengan PDF content
        """
        try:
            # Create response
            response = HttpResponse(content_type='application/pdf')
            
            if filename:
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
            else:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                response['Content-Disposition'] = f'attachment; filename="report_{timestamp}.pdf"'
            
            # Create PDF document
            buffer = io.BytesIO()
            doc = SimpleDocTemplate(
                buffer, 
                pagesize=A4,
                rightMargin=1*cm,
                leftMargin=1*cm,
                topMargin=1.5*cm,
                bottomMargin=1.5*cm
            )
            story = []
            
            # Add title
            story.append(Paragraph(title, self.title_style))
            story.append(Spacer(1, 0.2*inch))
            
            # Add date
            date_str = datetime.now().strftime('%d %B %Y, %H:%M:%S')
            story.append(Paragraph(f"Generated on: {date_str}", self.normal_style))
            story.append(Spacer(1, 0.3*inch))
            
            # Create cards for each item
            if data:
                for idx, item in enumerate(data):
                    # Create card content
                    card_elements = []
                    
                    # Card container table (2 columns: image and info)
                    card_data = []
                    
                    # Left column: Image - Fixed width, height adjusts to aspect ratio
                    image_path = item.get(image_field)
                    fixed_image_width = 2.8 * inch  # Fixed width untuk semua gambar
                    max_image_height = 3.5 * inch   # Max height untuk mencegah terlalu tinggi
                    
                    image_cell = ['']
                    if image_path and os.path.exists(image_path):
                        try:
                            # Resize image dengan lebar tetap
                            img = PILImage.open(image_path)
                            original_width, original_height = img.size
                            aspect_ratio = original_width / original_height
                            
                            # Hitung tinggi berdasarkan lebar tetap
                            calculated_height = fixed_image_width / aspect_ratio
                            
                            # Jika terlalu tinggi, batasi tinggi dan sesuaikan lebar
                            if calculated_height > max_image_height:
                                calculated_height = max_image_height
                                calculated_width = calculated_height * aspect_ratio
                            else:
                                calculated_width = fixed_image_width
                            
                            # Resize image
                            img_resized = img.resize(
                                (int(calculated_width), int(calculated_height)), 
                                PILImage.Resampling.LANCZOS
                            )
                            
                            # Save resized image to buffer
                            img_buffer = io.BytesIO()
                            img_resized.save(img_buffer, format='JPEG', quality=85)
                            img_buffer.seek(0)
                            
                            # Create ReportLab Image dengan ukuran yang sudah dihitung
                            rl_img = RLImage(
                                img_buffer, 
                                width=calculated_width, 
                                height=calculated_height
                            )
                            image_cell = [rl_img]
                        except Exception as e:
                            logger.warning(f"Error processing image {image_path}: {e}")
                            # Placeholder dengan ukuran tetap
                            image_cell = [Paragraph(
                                "No Image<br/>Available", 
                                ParagraphStyle(
                                    'NoImage',
                                    parent=self.normal_style,
                                    fontSize=10,
                                    textColor=colors.HexColor('#95a5a6'),
                                    alignment=1  # Center
                                )
                            )]
                    else:
                        # Placeholder dengan ukuran tetap
                        image_cell = [Paragraph(
                            "No Image<br/>Available", 
                            ParagraphStyle(
                                'NoImage',
                                parent=self.normal_style,
                                fontSize=10,
                                textColor=colors.HexColor('#95a5a6'),
                                alignment=1  # Center
                            )
                        )]
                    
                    # Right column: Bird information
                    bird_name = item.get('Bird Name', 'N/A')
                    scientific_name = item.get('Scientific Name', 'N/A')
                    family = item.get('Family', 'N/A')
                    description = item.get('Description', 'N/A')
                    habitat = item.get('Habitat', 'N/A')
                    
                    # Truncate long text for PDF
                    if description and len(description) > 200:
                        description = description[:200] + '...'
                    if habitat and len(habitat) > 200:
                        habitat = habitat[:200] + '...'
                    
                    # Create info as a single paragraph with HTML formatting
                    info_text = f"""
                    <b>{bird_name}</b><br/>
                    <i>{scientific_name}</i><br/><br/>
                    <b>Family:</b> {family}<br/><br/>
                    <b>Description:</b><br/>{description}<br/><br/>
                    <b>Habitat:</b><br/>{habitat}
                    """
                    
                    info_paragraph = Paragraph(info_text, ParagraphStyle(
                        'CardInfo',
                        parent=self.normal_style,
                        fontSize=10,
                        leading=12,
                        leftIndent=0,
                        rightIndent=0
                    ))
                    
                    # Create card table (2 columns: image and info)
                    # Lebar kolom: gambar 3 inch (dengan padding), info 4.8 inch
                    card_table = Table(
                        [[image_cell[0], info_paragraph]],
                        colWidths=[3.0*inch, 4.8*inch]
                    )
                    
                    # Style the card dengan spacing yang lebih baik
                    card_table.setStyle(TableStyle([
                        # Border around card
                        ('BOX', (0, 0), (-1, -1), 1, colors.HexColor('#bdc3c7')),
                        ('BACKGROUND', (0, 0), (-1, -1), colors.white),
                        ('VALIGN', (0, 0), (0, -1), 'TOP'),
                        ('VALIGN', (1, 0), (1, -1), 'TOP'),
                        # Padding untuk kolom gambar
                        ('LEFTPADDING', (0, 0), (0, -1), 12),
                        ('RIGHTPADDING', (0, 0), (0, -1), 12),
                        ('TOPPADDING', (0, 0), (0, -1), 12),
                        ('BOTTOMPADDING', (0, 0), (0, -1), 12),
                        # Padding untuk kolom info dengan spacing lebih besar dari kiri
                        ('LEFTPADDING', (1, 0), (1, -1), 20),  # Lebih besar untuk spacing dari gambar
                        ('RIGHTPADDING', (1, 0), (1, -1), 12),
                        ('TOPPADDING', (1, 0), (1, -1), 12),
                        ('BOTTOMPADDING', (1, 0), (1, -1), 12),
                    ]))
                    
                    # Add card to story
                    story.append(KeepTogether(card_table))
                    story.append(Spacer(1, 0.3*inch))
                    
                    # Page break every 2 cards
                    if (idx + 1) % 2 == 0:
                        story.append(Spacer(1, 0.2*inch))
            else:
                story.append(Paragraph("No data available", self.normal_style))
            
            # Build PDF
            doc.build(story)
            
            # Get PDF content
            pdf = buffer.getvalue()
            buffer.close()
            
            response.write(pdf)
            return response
            
        except Exception as e:
            logger.error(f"Error generating PDF cards: {e}")
            raise


class ExcelExporter:
    """
    Utility class untuk export data ke Excel menggunakan openpyxl
    """
    
    def __init__(self):
        self.workbook = None
        self.worksheet = None
    
    def export_to_excel(self, data, filename=None, title="Report", headers=None, sheet_name="Sheet1"):
        """
        Export data ke Excel
        
        Args:
            data: List of dictionaries atau list of lists
            filename: Nama file (optional)
            title: Judul report
            headers: List header untuk tabel (optional)
            sheet_name: Nama sheet
        
        Returns:
            HttpResponse dengan Excel content
        """
        try:
            # Create workbook
            self.workbook = Workbook()
            self.worksheet = self.workbook.active
            self.worksheet.title = sheet_name
            
            # Setup styles
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=12)
            title_font = Font(bold=True, size=14)
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            center_alignment = Alignment(horizontal='center', vertical='center')
            
            current_row = 1
            
            # Add title
            if title:
                self.worksheet.merge_cells(f'A{current_row}:{get_column_letter(len(headers) if headers else 1)}{current_row}')
                title_cell = self.worksheet[f'A{current_row}']
                title_cell.value = title
                title_cell.font = title_font
                title_cell.alignment = center_alignment
                current_row += 1
            
            # Add date
            date_str = datetime.now().strftime('%d %B %Y, %H:%M:%S')
            if headers:
                self.worksheet.merge_cells(f'A{current_row}:{get_column_letter(len(headers))}{current_row}')
            date_cell = self.worksheet[f'A{current_row}']
            date_cell.value = f"Generated on: {date_str}"
            date_cell.alignment = Alignment(horizontal='left')
            current_row += 1
            
            # Add empty row
            current_row += 1
            
            # Prepare headers
            if headers:
                header_row = current_row
                for col_idx, header in enumerate(headers, start=1):
                    cell = self.worksheet.cell(row=current_row, column=col_idx)
                    cell.value = str(header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = border
                current_row += 1
            elif data and isinstance(data[0], dict):
                # Auto-generate headers from first dict keys
                headers = list(data[0].keys())
                header_row = current_row
                for col_idx, header in enumerate(headers, start=1):
                    cell = self.worksheet.cell(row=current_row, column=col_idx)
                    cell.value = str(header)
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.alignment = center_alignment
                    cell.border = border
                current_row += 1
            
            # Add data rows
            if data:
                if isinstance(data[0], dict):
                    # Data is list of dictionaries
                    for row_data in data:
                        for col_idx, header in enumerate(headers, start=1):
                            cell = self.worksheet.cell(row=current_row, column=col_idx)
                            cell.value = str(row_data.get(header, ''))
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.border = border
                        current_row += 1
                else:
                    # Data is list of lists
                    for row_data in data:
                        for col_idx, value in enumerate(row_data, start=1):
                            cell = self.worksheet.cell(row=current_row, column=col_idx)
                            cell.value = str(value)
                            cell.alignment = Alignment(horizontal='left', vertical='center')
                            cell.border = border
                        current_row += 1
            
            # Auto-adjust column widths
            if headers:
                for col_idx, header in enumerate(headers, start=1):
                    column_letter = get_column_letter(col_idx)
                    max_length = len(str(header))
                    
                    # Check data rows for max length
                    for row_idx in range(header_row + 1, current_row):
                        cell_value = self.worksheet.cell(row=row_idx, column=col_idx).value
                        if cell_value:
                            max_length = max(max_length, len(str(cell_value)))
                    
                    # Set column width (add some padding)
                    adjusted_width = min(max_length + 2, 50)
                    self.worksheet.column_dimensions[column_letter].width = adjusted_width
            
            # Create response
            response = HttpResponse(
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            
            if filename:
                response['Content-Disposition'] = f'attachment; filename="{filename}"'
            else:
                timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
                response['Content-Disposition'] = f'attachment; filename="report_{timestamp}.xlsx"'
            
            # Save workbook to response
            self.workbook.save(response)
            return response
            
        except Exception as e:
            logger.error(f"Error generating Excel: {e}")
            raise
        finally:
            if self.workbook:
                self.workbook.close()


# Global instances
pdf_exporter = PDFExporter()
excel_exporter = ExcelExporter()

